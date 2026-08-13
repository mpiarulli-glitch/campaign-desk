#!/usr/bin/env python3
"""Push Krak Boba Corporate's spreadsheet calendar into Campaign Desk.

Reads the local editorial workbook, logs into the Campaign Desk API, removes
previous rows from this import marker, then creates planned calendar sends.
"""

from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime
from http.cookiejar import CookieJar
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import HTTPCookieProcessor, Request, build_opener

from openpyxl import load_workbook


CAMPAIGN_DESK_ROOT = Path(__file__).resolve().parents[2]
WORKSPACE = Path(__file__).resolve().parents[3]
ENV_PATH = CAMPAIGN_DESK_ROOT / "campaign-desk-sync" / ".env"
WORKBOOK = (
    WORKSPACE
    / "clients"
    / "Krak Boba Corporate"
    / "marketing"
    / "campaigns"
    / "Krak Boba Editorial Calendar 2026.xlsx"
)

CLIENT_NAME = "Krak Boba Corporate"
MARKER = "import:krak-boba-corporate-editorial-2026"
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def read_env() -> dict[str, str]:
    env = {}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def asset_type(channel: str, platform: str) -> str:
    text = f"{channel} {platform}".lower()
    if "email" in text:
        return "email_campaign"
    if "sms" in text:
        return "crm_automation"
    if "blog" in text:
        return "blog_post"
    if "video" in text or "tiktok" in text or "reel" in text:
        return "social_video_carousel"
    if "instagram" in text or "facebook" in text or "linkedin" in text:
        return "social_post"
    return ""


def subject_from_hook(channel: str, hook: str) -> str:
    if channel.lower() != "email":
        return ""
    return re.sub(r"^subject:\s*", "", hook, flags=re.I).strip()


def load_rows() -> list[dict[str, str]]:
    wb = load_workbook(WORKBOOK, data_only=True)
    rows: list[dict[str, str]] = []
    for ws in wb.worksheets:
        if ws.title == "Q2 2026":
            continue
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            date, _day, channel, platform, campaign, name, desc, hook, cta, segment, tie = row[:11]
            if not any(clean(v) for v in (date, channel, platform, campaign, name, desc, hook, cta, segment, tie)):
                continue
            send_date = clean(date)
            if not DATE_RE.match(send_date):
                raise ValueError(f"{ws.title} row {row_num}: invalid date {send_date!r}")
            channel_s = clean(channel)
            platform_s = clean(platform)
            campaign_s = clean(campaign)
            name_s = clean(name)
            title = f"{campaign_s}: {name_s}" if campaign_s and name_s else campaign_s or name_s
            if not title:
                raise ValueError(f"{ws.title} row {row_num}: missing campaign/name")
            description = clean(desc)
            hook_s = clean(hook)
            cta_s = clean(cta)
            tie_s = clean(tie)
            note_parts = [
                f"Source workbook: {WORKBOOK.name}, {ws.title} row {row_num}",
                f"Channel: {channel_s}" if channel_s else "",
                f"Hook/message: {hook_s}" if hook_s else "",
                f"CTA: {cta_s}" if cta_s else "",
                f"Seasonal tie-in: {tie_s}" if tie_s else "",
                MARKER,
            ]
            rows.append(
                {
                    "title": title,
                    "sendDate": send_date,
                    "sendTime": "",
                    "status": "planned",
                    "platform": platform_s or channel_s,
                    "assetType": asset_type(channel_s, platform_s),
                    "audience": clean(segment),
                    "purpose": description,
                    "offer": cta_s,
                    "subject": subject_from_hook(channel_s, hook_s),
                    "previewText": "",
                    "note": "\n".join(part for part in note_parts if part),
                }
            )
    return rows


class Desk:
    def __init__(self, base: str, password: str):
        self.base = base.rstrip("/")
        self.password = password
        self.opener = build_opener(HTTPCookieProcessor(CookieJar()))

    def request(self, method: str, path: str, payload: dict | None = None):
        data = None
        headers = {}
        if payload is not None:
            data = json.dumps(payload).encode()
            headers["Content-Type"] = "application/json"
        req = Request(f"{self.base}{path}", data=data, headers=headers, method=method)
        try:
            with self.opener.open(req) as res:
                body = res.read().decode()
                return json.loads(body) if body else {}
        except HTTPError as exc:
            detail = exc.read().decode(errors="replace")
            raise RuntimeError(f"{method} {path} failed: {exc.code} {detail}") from exc

    def login(self):
        self.request("POST", "/api/auth", {"password": self.password})

    def clients(self):
        return self.request("GET", "/api/revenue/clients").get("clients", [])

    def ensure_client(self):
        clients = self.clients()
        found = next((c for c in clients if c.get("name") == CLIENT_NAME), None)
        if found:
            return found
        return self.request(
            "POST",
            "/api/revenue/clients",
            {"name": CLIENT_NAME, "businessModel": "hospitality"},
        )["client"]

    def list_sends(self, start: str, end: str):
        query = urlencode({"start": start, "end": end, "all": "1"})
        return self.request("GET", f"/api/calendar?{query}").get("sends", [])

    def delete_send(self, send_id: str):
        self.request("DELETE", f"/api/calendar/{send_id}")

    def create_send(self, payload: dict):
        return self.request("POST", "/api/calendar", payload)


def main() -> int:
    dry_run = "--dry-run" in sys.argv
    env = read_env()
    base = os.environ.get("CAMPAIGN_DESK_URL") or env.get("CAMPAIGN_DESK_URL")
    password = os.environ.get("CAMPAIGN_DESK_PASSWORD") or env.get("CAMPAIGN_DESK_PASSWORD")
    if not base or not password:
        raise SystemExit("Missing CAMPAIGN_DESK_URL or CAMPAIGN_DESK_PASSWORD")

    rows = load_rows()
    if not rows:
        raise SystemExit("No calendar rows found")
    start = min(row["sendDate"] for row in rows)
    end = max(row["sendDate"] for row in rows)
    print(f"Loaded {len(rows)} rows from {WORKBOOK}")
    print(f"Window: {start} to {end}")

    desk = Desk(base, password)
    desk.login()
    client = desk.ensure_client()
    print(f"Campaign Desk client: {client['name']} ({client['id']})")

    existing = desk.list_sends(start, end)
    marked = [
        send
        for send in existing
        if send.get("client_id") == client["id"] and MARKER in (send.get("note") or "")
    ]
    print(f"Existing marked rows to replace: {len(marked)}")
    if dry_run:
        by_type = {}
        for row in rows:
            key = row["assetType"] or "unmapped"
            by_type[key] = by_type.get(key, 0) + 1
        print("Dry run only. Asset type counts:")
        for key, value in sorted(by_type.items()):
            print(f"  {key}: {value}")
        print("First five rows:")
        for row in rows[:5]:
            print(f"  {row['sendDate']} | {row['assetType']} | {row['title']}")
        return 0

    for send in marked:
        desk.delete_send(send["id"])
    created = 0
    for row in rows:
        payload = {"clientId": client["id"], **row}
        desk.create_send(payload)
        created += 1
    print(f"Deleted {len(marked)} marked rows and created {created} planned sends.")
    print(f"Admin calendar: {base}/admin/calendar")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
