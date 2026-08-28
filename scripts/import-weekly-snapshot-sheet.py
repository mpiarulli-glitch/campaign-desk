#!/usr/bin/env python3
"""
Import Weekly Client Snapshot.xlsx into live Campaign Desk snapshots.

Only the 15 snapshot-roster accounts. Former-client tabs and Template are
skipped. Krak Boba Corporate only (not Temecula). Roster accounts with no
tab stay empty.

Column A is the deliverable name on the standard template (Wk1 starts in
column B). On roadmap sheets where Wk1 starts later, A is a merged category
and B is the line-item name.

Usage:
  python3 scripts/import-weekly-snapshot-sheet.py            # dry run
  python3 scripts/import-weekly-snapshot-sheet.py --commit   # push to hub
"""

from __future__ import annotations

import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

SHEET_PATH = Path(
    os.environ.get(
        "SNAPSHOT_XLSX",
        str(Path.home() / "Downloads" / "Weekly Client Snapshot.xlsx"),
    )
)
URL = os.environ.get("CAMPAIGN_DESK_URL", "https://hub.marketingempiregroup.com")
PASSWORD = os.environ.get("CAMPAIGN_DESK_PASSWORD", "Marketingeg1!")
# Owner login is 2FA-gated; named admins without TOTP can still write snapshots.
LOGIN_ACCOUNTS = [
    s
    for s in os.environ.get("CAMPAIGN_DESK_ACCOUNT", "kyle_onstott,luis_romero,michael").split(",")
    if s.strip()
]
COMMIT = "--commit" in sys.argv
DEFAULT_YEAR = 2026

# Sheet tab -> allowlist phrase (must match isSnapshotAllowlisted on the hub).
ROSTER_SHEETS = {
    "BetterLife Coach": "Betterlife Coach",
    "Pacific Coast Generation": "Pacific Coast Generation",
    "The HR Innovator Group": "HR Innovator Group",
    "Hendos Barrel House": "Hendo's Barrel House",
    "Ecoworkz": "Ecoworkz",
    "CISCo Restaurant + Bar": "CISCo Restauraunt + Bar",
    "Pipe It RIght": "Pipe It Right",
    "CIPO Cloud Software": "CIPO Cloud Software",
    "Guardian Plumbers": "Guardian Plumbers",
    "12 Volt Power": "12 Volt Power",
    "Kentina Hospitality": "Kentina Hospitality",
    "Krak Boba (Corporate)": "Krak Boba Corporate",
}

ROSTER_WITHOUT_TAB = [
    "Looda House Pawn",
    "Our Watch / tim thompson",
    "Vitatherapy Wellness Spa",
]

MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

STATUS_MAP = {
    "completed": "completed",
    "complete": "completed",
    "share & approved": "approved",
    "shared & approved": "approved",
    "shared and approved": "approved",
    "approved": "approved",
    "shared & not yet approved": "shared",
    "shared and not yet approved": "shared",
    "shared - not yet approved": "shared",
    "not yet approved": "shared",
    "not shared": "shared",
    "shared": "shared",
    "shared — awaiting approval": "shared",
    "in progress": "in_progress",
    "ongoing": "in_progress",
    "assigned": "not_started",
    "on hold": "not_started",
    "not started": "not_started",
    "scheduled": "scheduled",
    "sent for approval": "sent_for_approval",
    "sent": "sent_for_approval",
    "canceled": "canceled",
    "cancelled": "canceled",
}

SECTION_RE = re.compile(
    r"^(strategy\s*&\s*planning|reviews?\s*&\s*reputation|content|seo\s*&\s*visibility|"
    r"crm,?\s*email\s*&\s*automation|website,?\s*tracking\s*&\s*ai|"
    r"social media strategy\s*&\s*management|content creation|email\s*&\s*outreach|"
    r"visibility\s*&\s*listings|visibility\s*\(|conversion\s*\(|nurturing\s*&\s*automation|"
    r"2026 q1.*|à la carte requests|a la carte requests|"
    r"insert contract deliverables)$",
    re.I,
)
STOP_TOKENS = {
    "a",
    "an",
    "the",
    "and",
    "or",
    "of",
    "to",
    "for",
    "per",
    "up",
    "avg",
    "total",
    "hours",
    "hour",
    "hrs",
    "hr",
    "month",
    "monthly",
    "weekly",
    "week",
    "daily",
    "ongoing",
    "one",
    "time",
    "updated",
    "management",
    "media",
}
PLACEHOLDER_RE = re.compile(r"^[\s\-–—•\u200b\u00a0]*$")


def text(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return re.sub(r"[\s\u00a0\u200b]+", " ", str(v)).replace("\n", " ").strip()


def is_placeholder(s: str) -> bool:
    return not s or bool(PLACEHOLDER_RE.match(s))


def norm_name(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def parse_status(label: str) -> str | None:
    n = re.sub(r"\s+", " ", label.lower()).strip()
    n = n.replace("—", "-").replace("–", "-")
    if n in STATUS_MAP:
        return STATUS_MAP[n]
    for k, v in STATUS_MAP.items():
        if n.startswith(k) or k in n:
            return v
    return None


def cadence_for(name: str) -> tuple[str, str, str]:
    """Return (cadence label, kind, cadence_unit)."""
    n = name.lower()
    if "one time" in n or "one-time" in n or "1x)" in n and "set up" in n:
        return ("One-time", "one_time", "monthly")
    if "quarterly" in n:
        return ("Quarterly", "recurring", "quarterly")
    if "weekly" in n or "/wk" in n or "posts/week" in n or "x week" in n:
        return ("Weekly", "recurring", "weekly")
    if "monthly" in n or "/month" in n or "per month" in n or "x/month" in n or "x month" in n:
        return ("Monthly", "recurring", "monthly")
    return ("", "recurring", "monthly")


def first_monday(year: int, month: int) -> date:
    d = date(year, month, 1)
    add = 0 if d.weekday() == 0 else (7 - d.weekday()) % 7
    return d + timedelta(days=add)


def week_monday(year: int, month: int, week_n: int) -> str:
    d = first_monday(year, month) + timedelta(days=(week_n - 1) * 7)
    return d.isoformat()


class CookieJar:
    def __init__(self):
        self.cookie = ""

    def api(self, method: str, path: str, body=None):
        data = None if body is None else json.dumps(body).encode()
        headers = {"Content-Type": "application/json"}
        if self.cookie:
            headers["Cookie"] = self.cookie
        req = urllib.request.Request(URL + path, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=60) as res:
                sc = res.headers.get("Set-Cookie")
                if sc:
                    self.cookie = sc.split(";")[0]
                raw = res.read().decode()
        except urllib.error.HTTPError as e:
            raw = e.read().decode(errors="replace")
            raise RuntimeError(f"{method} {path} -> {e.code} {raw[:400]}") from e
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return raw


def match_account(accounts, phrase: str):
    want = norm_name(phrase)
    for a in accounts:
        n = norm_name(a["name"])
        if n == want or want in n or n in want:
            return a
    # token overlap
    pt = [t for t in re.split(r"[^a-z0-9]+", phrase.lower()) if len(t) > 2]
    for a in accounts:
        ct = set(re.split(r"[^a-z0-9]+", a["name"].lower()))
        if pt and all(any(t in c or c in t for c in ct) for t in pt):
            return a
    return None


def fold_ai(s: str) -> str:
    return re.sub(r"\bal\b", "ai", s.lower())


TOKEN_ALIASES = {
    "blogging": "blog",
    "graphics": "graphic",
    "influencers": "influencer",
    "productions": "production",
    "posts": "post",
}


def name_tokens(s: str) -> set[str]:
    parts = re.findall(r"[a-z0-9]+", fold_ai(s))
    out = set()
    for p in parts:
        p = TOKEN_ALIASES.get(p, p)
        if p in STOP_TOKENS:
            continue
        if len(p) > 1 or p.isdigit():
            out.add(p)
    return out


def hours_in(s: str) -> str | None:
    m = re.search(r"(\d+)\s*(?:hours?|hrs?)", s.lower())
    return m.group(1) if m else None


def match_deliverable(remote, name: str):
    n = norm_name(fold_ai(name))
    by = {norm_name(fold_ai(d["name"])): d for d in remote}
    if n in by:
        return by[n]
    key = n[:14]
    prefix_hits = []
    for d in remote:
        dn = norm_name(fold_ai(d["name"]))
        if key and (dn.startswith(key) or n.startswith(dn[:14])):
            prefix_hits.append(d)
    if len(prefix_hits) == 1:
        return prefix_hits[0]
    want = name_tokens(name)
    want_hrs = hours_in(name)
    scored = []
    distinctive = {
        "email",
        "seo",
        "blog",
        "blogging",
        "social",
        "chatbot",
        "gbp",
        "google",
        "influencer",
        "influencers",
        "production",
        "productions",
        "graphic",
        "graphics",
        "leads",
        "linkedin",
        "meta",
        "landing",
    }
    for d in remote:
        got = name_tokens(d["name"])
        inter = want & got
        if want_hrs:
            got_hrs = hours_in(d["name"])
            if got_hrs and got_hrs != want_hrs:
                continue
        if len(inter) >= 3 or (len(inter) >= 2 and inter & distinctive):
            scored.append((len(inter), d))
    scored.sort(key=lambda x: -x[0])
    if scored and (len(scored) == 1 or scored[0][0] > scored[1][0]):
        return scored[0][1]
    return None


def vertical_a_merge(ws, row: int):
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1 and mr.max_col == 1 and mr.min_row <= row <= mr.max_row:
            return mr.max_row - mr.min_row + 1, mr.min_row
    return 1, row


def parse_sheet(ws):
    max_c = min(ws.max_column or 1, 120)
    max_r = min(ws.max_row or 1, 200)

    def val(r, c):
        return text(ws.cell(r, c).value)

    month_row = wk_row = None
    for r in range(1, 8):
        months = 0
        wks = 0
        for c in range(1, max_c + 1):
            t = val(r, c)
            if re.search(r"In Progress:\s*[A-Za-z]+", t, re.I) or (
                re.match(
                    r"^(January|February|March|April|May|June|July|August|September|October|November|December)\b",
                    t,
                    re.I,
                )
                and "launch" not in t.lower()
            ):
                months += 1
            if re.match(r"^Wk\d$", t):
                wks += 1
        if months >= 2 and month_row is None:
            month_row = r
        if wks >= 3 and wk_row is None:
            wk_row = r
    if month_row is None or wk_row is None:
        return {"error": "Could not find month/week headers", "deliverables": [], "entries": []}

    # Column map: each col -> {month, year, kind}
    col_info = []
    last_month = None
    last_year = DEFAULT_YEAR
    last_idx = None
    first_wk_col = None
    for c in range(1, max_c + 1):
        mh = val(month_row, c)
        m = re.search(
            r"(?:In Progress:\s*)?(January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(\d{4}))?",
            mh,
            re.I,
        )
        if m and ("progress" in mh.lower() or mh.lower().startswith(m.group(1).lower())):
            month = m.group(1).title()
            idx = MONTHS[month.lower()]
            year = int(m.group(2)) if m.group(2) else last_year
            if last_idx is not None and idx < last_idx:
                year = last_year + 1
            last_month, last_year, last_idx = month, year, idx
        sub = val(wk_row, c)
        kind = None
        wk = re.match(r"^Wk(\d)$", sub)
        if wk:
            kind = ("week", int(wk.group(1)))
            if first_wk_col is None:
                first_wk_col = c
        elif re.match(r"^Status", sub, re.I) or (re.match(r"^Status", mh, re.I) and not wk):
            if re.match(r"^Status", sub, re.I) or (not sub and re.match(r"^Status", mh, re.I)):
                kind = ("status", None)
        elif re.match(r"^Next Steps", sub, re.I) or re.match(r"^Next Steps", mh, re.I):
            kind = ("next", None)
        elif re.match(r"^Notes", sub, re.I) or re.match(r"^Notes", mh, re.I):
            kind = ("notes", None)
        # Status/Next/Notes often live on the month row while the week row is blank
        if kind is None:
            if re.match(r"^Status", mh, re.I):
                kind = ("status", None)
            elif re.match(r"^Next Steps", mh, re.I):
                kind = ("next", None)
            elif re.match(r"^Notes", mh, re.I):
                kind = ("notes", None)
        col_info.append({"month": last_month, "year": last_year, "kind": kind, "col": c})

    if first_wk_col is None:
        return {"error": "No Wk columns", "deliverables": [], "entries": []}

    name_in_b = first_wk_col >= 3
    name_col = 2 if name_in_b else 1

    # Group columns by month block
    blocks = []  # {month, year, weeks:{n:col}, status, next, notes}
    current = None
    seen_notes = False
    for info in col_info:
        k = info["kind"]
        if k and k[0] == "week" and info["month"]:
            key = (info["month"], info["year"])
            if current is None or (current["month"], current["year"]) != key:
                current = {
                    "month": info["month"],
                    "year": info["year"],
                    "weeks": {},
                    "status": None,
                    "next": None,
                    "notes": None,
                }
                blocks.append(current)
            if k[1] not in current["weeks"]:
                current["weeks"][k[1]] = info["col"]
            seen_notes = False
        elif k and current is not None:
            if k[0] == "status" and current["status"] is None:
                current["status"] = info["col"]
            elif k[0] == "next" and current["next"] is None:
                current["next"] = info["col"]
            elif k[0] == "notes" and current["notes"] is None:
                current["notes"] = info["col"]
                seen_notes = True

    last_category = ""
    last_name = ""
    deliverables = []
    empty_run = 0
    for r in range(wk_row + 1, max_r + 1):
        a = val(r, 1)
        b = val(r, 2)
        height, _top = vertical_a_merge(ws, r)
        if name_in_b:
            name = b if b and not re.match(r"^Wk\d$", b) and b.lower() not in ("notes", "date", "status") else ""
            if a:
                last_category = a
            category = last_category
            if not name and a and height == 1 and not SECTION_RE.match(a):
                name = a
        else:
            name = a
            category = last_category

        def row_has_work():
            for info in col_info:
                if not info["kind"] or info["kind"][0] not in ("week", "status", "next", "notes"):
                    continue
                t = val(r, info["col"])
                if t and not is_placeholder(t) and not re.match(r"^Wk\d$", t):
                    if re.match(r"^(status|next steps|notes)$", t, re.I):
                        continue
                    return True
            return False

        wrap = False
        if name and last_name and deliverables and not row_has_work():
            if name[0].islower() or name[:1] in ",)&":
                wrap = True
            elif last_name.rstrip().endswith(("+", "-", ",", "&", "/")):
                wrap = True
        if wrap:
            deliverables[-1]["name"] = (deliverables[-1]["name"] + " " + name).strip()
            last_name = deliverables[-1]["name"]
            empty_run = 0
            continue

        if not name:
            empty_run += 1
            if empty_run >= 8 and deliverables:
                break
            continue
        empty_run = 0

        if SECTION_RE.match(name) or (name.isupper() and len(name) < 48 and "&" in name):
            last_category = name
            continue

        if re.match(r"^insert contract", name, re.I):
            continue
        if re.match(r"^à la carte requests$|^a la carte requests$", name, re.I) and (not b or b.lower() == "date"):
            last_category = name
            continue

        last_name = name
        cadence, kind, unit = cadence_for(name)
        row_entries = []
        for blk in blocks:
            status_txt = val(r, blk["status"]) if blk["status"] else ""
            next_txt = val(r, blk["next"]) if blk["next"] else ""
            notes_txt = val(r, blk["notes"]) if blk["notes"] else ""
            month_status = parse_status(status_txt)
            if is_placeholder(next_txt):
                next_txt = ""
            if is_placeholder(notes_txt):
                notes_txt = ""
            if next_txt and parse_status(next_txt) and len(next_txt) < 40:
                next_txt = ""
            weeks_with = []
            for n in range(1, 6):
                col = blk["weeks"].get(n)
                if not col:
                    continue
                raw = val(r, col)
                if is_placeholder(raw):
                    continue
                chip = parse_status(raw)
                work = raw
                if chip and len(raw) < 48:
                    work = ""
                if work or chip:
                    weeks_with.append({"n": n, "work": work, "chip": chip})
            month_entries = []
            if weeks_with:
                last_i = len(weeks_with) - 1
                for i, w in enumerate(weeks_with):
                    month_entries.append(
                        {
                            "weekStart": week_monday(blk["year"], MONTHS[blk["month"].lower()], w["n"]),
                            "status": w["chip"] or month_status or "not_started",
                            "workDone": w["work"],
                            "nextSteps": next_txt if i == last_i else "",
                            "notes": notes_txt if i == last_i else "",
                            "label": f"{blk['month']} {blk['year']} Wk{w['n']}",
                        }
                    )
            elif month_status or next_txt or notes_txt:
                month_entries.append(
                    {
                        "weekStart": week_monday(blk["year"], MONTHS[blk["month"].lower()], 1),
                        "status": month_status or "not_started",
                        "workDone": "",
                        "nextSteps": next_txt,
                        "notes": notes_txt,
                        "label": f"{blk['month']} {blk['year']} (month)",
                    }
                )
            row_entries.extend(e for e in month_entries if e["weekStart"].startswith("2026-"))

        deliverables.append(
            {
                "row": r,
                "name": name,
                "category": category if category != name else "",
                "cadence": cadence,
                "kind": kind,
                "cadenceUnit": unit,
                "entries": row_entries,
            }
        )

    return {
        "error": None,
        "name_in_b": name_in_b,
        "first_wk_col": first_wk_col,
        "blocks": [
            {
                "month": b["month"],
                "year": b["year"],
                "weeks": sorted(b["weeks"]),
                "status": b["status"],
                "next": b["next"],
                "notes": b["notes"],
            }
            for b in blocks
        ],
        "deliverables": deliverables,
        "entries": [e for d in deliverables for e in d["entries"]],
    }


def main():
    if not SHEET_PATH.exists():
        print(f"Missing spreadsheet: {SHEET_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {SHEET_PATH} …")
    wb = load_workbook(SHEET_PATH, data_only=True, read_only=False)
    parsed = {}
    for tab in ROSTER_SHEETS:
        if tab not in wb.sheetnames:
            parsed[tab] = {"error": "Tab missing in workbook", "deliverables": [], "entries": []}
            continue
        parsed[tab] = parse_sheet(wb[tab])

    jar = CookieJar()
    logged_in = None
    for slug in LOGIN_ACCOUNTS:
        jar.cookie = ""
        try:
            login = jar.api("POST", "/api/auth", {"password": PASSWORD, "account": slug.strip()})
        except RuntimeError as e:
            print(f"  {slug}: {e}")
            continue
        if isinstance(login, dict) and login.get("needsTotp"):
            print(f"  {slug}: 2FA required, skipping")
            continue
        if isinstance(login, dict) and login.get("ok") and not login.get("needsTotp"):
            logged_in = slug.strip()
            break
    if not logged_in:
        print("Could not log in to the hub (all accounts failed or need 2FA).", file=sys.stderr)
        sys.exit(1)
    accounts = jar.api("GET", "/api/snapshot/accounts").get("accounts") or []
    print(f"Connected to {URL} as {logged_in}  ({len(accounts)} snapshot accounts)\n")

    planned = []
    print("Skipped (no tab): " + ", ".join(ROSTER_WITHOUT_TAB))
    print()

    for tab, phrase in ROSTER_SHEETS.items():
        info = parsed[tab]
        acct = match_account(accounts, phrase) or match_account(accounts, tab)
        print(f"=== {tab} ===")
        if info.get("error"):
            print(f"  PARSE ERROR: {info['error']}")
            continue
        print(
            f"  layout: name_col={'B' if info['name_in_b'] else 'A'}  "
            f"wk_start_col={info['first_wk_col']}  "
            f"months={[(b['month'], b['year']) for b in info['blocks']]}"
        )
        if not acct:
            print(f"  NO MATCHING ACCOUNT on hub for '{phrase}'")
            continue
        detail = jar.api("GET", f"/api/snapshot/accounts/{acct['id']}")
        remote = detail.get("deliverables") or []
        print(f"  hub: {acct['name']}  ({len(remote)} existing deliverables)")
        unmatched_create = []
        matched = 0
        entry_count = 0
        for d in info["deliverables"]:
            rem = match_deliverable(remote, d["name"])
            if rem:
                matched += 1
                did = rem["id"]
                action = "match"
            else:
                unmatched_create.append(d["name"])
                did = None
                action = "create"
            entry_count += len(d["entries"])
            planned.append(
                {
                    "accountId": acct["id"],
                    "accountName": acct["name"],
                    "deliverableId": did,
                    "create": d if did is None else None,
                    "name": d["name"],
                    "action": action,
                    "entries": d["entries"],
                }
            )
        print(
            f"  sheet deliverables={len(info['deliverables'])}  "
            f"matched={matched}  to_create={len(unmatched_create)}  "
            f"entries={entry_count}"
        )
        if unmatched_create:
            preview = unmatched_create[:8]
            extra = f" (+{len(unmatched_create)-8} more)" if len(unmatched_create) > 8 else ""
            print("  create: " + " | ".join(preview) + extra)

    total_entries = sum(len(p["entries"]) for p in planned)
    to_create = sum(1 for p in planned if p["action"] == "create")
    print(
        f"\nTotal: {len(planned)} deliverable rows, {to_create} new deliverables, "
        f"{total_entries} entries to {'PUSH' if COMMIT else 'push (dry run)'}"
    )

    if not COMMIT:
        print("\nDry run. Re-run with --commit to write to the hub.")
        return

    ok_d = fail_d = ok_e = fail_e = 0
    created_ids = {}
    for p in planned:
        did = p["deliverableId"]
        if p["action"] == "create":
            d = p["create"]
            try:
                res = jar.api(
                    "POST",
                    f"/api/snapshot/accounts/{p['accountId']}/deliverables",
                    {
                        "name": d["name"],
                        "category": d["category"],
                        "cadence": d["cadence"],
                        "kind": d["kind"],
                        "cadenceUnit": d["cadenceUnit"],
                    },
                )
                did = res["deliverable"]["id"]
                created_ids[p["name"]] = did
                ok_d += 1
            except Exception as e:
                fail_d += 1
                print(f"  FAIL create {p['accountName']} / {p['name']}: {e}")
                continue
        if not did:
            continue
        for i, e in enumerate(p["entries"], 1):
            try:
                jar.api(
                    "POST",
                    "/api/snapshot/entry",
                    {
                        "deliverableId": did,
                        "weekStart": e["weekStart"],
                        "status": e["status"],
                        "workDone": e["workDone"],
                        "nextSteps": e["nextSteps"],
                        "notes": e["notes"],
                    },
                )
                ok_e += 1
                if ok_e % 50 == 0:
                    print(f"  … {ok_e} entries")
            except Exception as err:
                fail_e += 1
                print(f"  FAIL entry {p['accountName']} {e['weekStart']}: {err}")
    print(f"\nCreated {ok_d} deliverables ({fail_d} failed). Pushed {ok_e} entries ({fail_e} failed).")


if __name__ == "__main__":
    main()
